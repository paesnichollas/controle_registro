from rest_framework import status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.contrib.auth.models import User, Group
from controle.serializers import UserSerializer
from controle.models import RegistroOS
import logging
from django.db.models import Q
from datetime import datetime, timedelta
import csv
from io import StringIO, BytesIO
from django.http import HttpResponse
import json
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class CustomTokenObtainPairView(TokenObtainPairView):
    """
    View customizada para obter token JWT com informações adicionais do usuário
    """
    
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        
        if response.status_code == 200:
            # Adicionar informações do usuário à resposta
            username = request.data.get('username')
            try:
                user = User.objects.get(username=username)
                user_serializer = UserSerializer(user)
                response.data['user'] = user_serializer.data
                
                # Adicionar informações dos grupos
                groups = [group.name for group in user.groups.all()]
                response.data['groups'] = groups
                
                logger.info(f"Login realizado com sucesso para o usuário: {username}")
                
            except User.DoesNotExist:
                logger.error(f"Usuário não encontrado: {username}")
        
        return response


class CustomTokenRefreshView(TokenRefreshView):
    """
    View customizada para refresh de token JWT
    """
    
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        
        if response.status_code == 200:
            logger.info("Token refresh realizado com sucesso")
        
        return response


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login(request):
    """
    Endpoint de login que retorna token JWT e informações do usuário
    """
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response({
            'error': 'Username e password são obrigatórios'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Primeiro, verificar se o usuário existe
    try:
        user_exists = User.objects.filter(username=username).exists()
    except Exception as e:
        logger.error(f"Erro ao verificar existência do usuário {username}: {str(e)}")
        return Response({
            'error': 'Erro interno do servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # Se o usuário não existe, retornar erro imediatamente
    if not user_exists:
        logger.warning(f"Tentativa de login com usuário inexistente: {username}")
        return Response({
            'error': 'Usuário não encontrado'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Autenticar usuário
    user = authenticate(username=username, password=password)
    
    if user is None:
        # Usuário existe mas senha está incorreta
        logger.warning(f"Senha incorreta para o usuário: {username}")
        return Response({
            'error': 'Senha incorreta'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    if not user.is_active:
        logger.warning(f"Tentativa de login de usuário inativo: {username}")
        return Response({
            'error': 'Conta de usuário desativada'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Verificar se o usuário tem grupos/permissões
    groups = [group.name for group in user.groups.all()]
    if not groups:
        logger.warning(f"Usuário {username} não possui grupos/permissões")
        return Response({
            'error': 'Usuário não possui permissão para acessar o sistema'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Gerar tokens JWT
    refresh = RefreshToken.for_user(user)
    access_token = refresh.access_token
    
    # Serializar dados do usuário
    user_serializer = UserSerializer(user)
    
    logger.info(f"Login realizado com sucesso para o usuário: {username}")
    
    return Response({
        'access': str(access_token),
        'refresh': str(refresh),
        'user': user_serializer.data,
        'groups': groups
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def logout(request):
    """
    Endpoint de logout que adiciona o token à blacklist
    """
    try:
        refresh_token = request.data.get('refresh')
        
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
        
        logger.info(f"Logout realizado para o usuário: {request.user.username}")
        
        return Response({
            'message': 'Logout realizado com sucesso'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Erro durante logout: {str(e)}")
        return Response({
            'error': 'Erro interno do servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def perfil(request):
    """
    Retorna informações do perfil do usuário logado
    """
    logger.warning(f"User ID: {request.user.id}, Username: {request.user.username}, Authorization: {request.META.get('HTTP_AUTHORIZATION')}")
    user_serializer = UserSerializer(request.user)
    groups = [group.name for group in request.user.groups.all()]
    
    return Response({
        'user': user_serializer.data,
        'groups': groups,
        'permissions': {
            'can_delete': 'Administrador' in groups,
            'can_edit_all': any(group in groups for group in ['Administrador', 'Superior']),
            'can_edit_financial': 'Cliente' not in groups,
            'is_admin': 'Administrador' in groups,
            'is_superior': 'Superior' in groups,
            'is_qualidade': 'Qualidade' in groups,
            'is_tecnico': 'Tecnico' in groups,
            'is_cliente': 'Cliente' in groups,
            'is_basico': 'Básico' in groups,
        }
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def alterar_senha(request):
    """
    Permite ao usuário alterar sua própria senha
    """
    senha_atual = request.data.get('senha_atual')
    nova_senha = request.data.get('nova_senha')
    confirmar_senha = request.data.get('confirmar_senha')
    
    if not all([senha_atual, nova_senha, confirmar_senha]):
        return Response({
            'error': 'Todos os campos são obrigatórios'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if nova_senha != confirmar_senha:
        return Response({
            'error': 'Nova senha e confirmação não coincidem'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if len(nova_senha) < 8:
        return Response({
            'error': 'A nova senha deve ter pelo menos 8 caracteres'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Verificar senha atual
    if not request.user.check_password(senha_atual):
        return Response({
            'error': 'Senha atual incorreta'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Alterar senha
    request.user.set_password(nova_senha)
    request.user.save()
    
    logger.info(f"Senha alterada para o usuário: {request.user.username}")
    
    return Response({
        'message': 'Senha alterada com sucesso'
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def verificar_token(request):
    """
    Verifica se o token JWT é válido
    """
    return Response({
        'valid': True,
        'user': request.user.username,
        'groups': [group.name for group in request.user.groups.all()]
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def listar_usuarios(request):
    """
    Lista todos os usuários (apenas para administradores e superiores)
    """
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem listar usuários.'
        }, status=status.HTTP_403_FORBIDDEN)
    usuarios = User.objects.filter(is_active=True)
    serializer = UserSerializer(usuarios, many=True)
    
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def criar_usuario(request):
    """
    Cria um novo usuário (apenas para administradores e superiores)
    """
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem criar usuários.'
        }, status=status.HTTP_403_FORBIDDEN)
    username = request.data.get('username')
    email = request.data.get('email')
    password = request.data.get('password')
    first_name = request.data.get('first_name', '')
    last_name = request.data.get('last_name', '')
    groups = request.data.get('groups', [])
    
    if not all([username, email, password]):
        return Response({
            'error': 'Username, email e password são obrigatórios'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if User.objects.filter(username=username).exists():
        return Response({
            'error': 'Username já existe'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if User.objects.filter(email=email).exists():
        return Response({
            'error': 'Email já existe'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Criar usuário
    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name
    )
    
    # Adicionar aos grupos
    for group_name in groups:
        try:
            group = Group.objects.get(name=group_name)
            user.groups.add(group)
        except Group.DoesNotExist:
            logger.warning(f"Grupo não encontrado: {group_name}")
    
    logger.info(f"Usuário criado: {username} por {request.user.username}")
    
    serializer = UserSerializer(user)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def listar_grupos(request):
    """
    Lista todos os grupos disponíveis
    """
    grupos = Group.objects.all()
    return Response([{'id': g.id, 'name': g.name} for g in grupos])


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def verificar_admin(request):
    """
    Verifica se o usuário é administrador ou superior
    """
    is_admin = request.user.groups.filter(name='Administrador').exists()
    is_superior = request.user.groups.filter(name='Superior').exists()
    return Response({
        'is_admin': is_admin,
        'is_superior': is_superior,
        'can_manage': is_admin or is_superior
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def relatorios_lista_registros(request):
    """
    Lista registros para relatórios (apenas para administradores e superiores)
    """
    logger.info(f"Relatórios: Usuário {request.user.username} tentando acessar relatórios")
    
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        logger.warning(f"Relatórios: Acesso negado para usuário {request.user.username}")
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem acessar relatórios.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Parâmetros de filtro
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    cliente = request.GET.get('cliente')
    status_os = request.GET.get('status_os')
    
    logger.info(f"Relatórios: Filtros - data_inicio={data_inicio}, data_fim={data_fim}, cliente={cliente}, status_os={status_os}")
    
    # Query base
    registros = RegistroOS.objects.select_related(
        'nome_cliente', 'status_os', 'usuario'
    ).all()
    
    # Aplicar filtros
    if data_inicio:
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            registros = registros.filter(created_at__gte=data_inicio)
        except ValueError:
            logger.warning(f"Relatórios: Data início inválida: {data_inicio}")
            pass
    
    if data_fim:
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            registros = registros.filter(created_at__lt=data_fim)
        except ValueError:
            logger.warning(f"Relatórios: Data fim inválida: {data_fim}")
            pass
    
    if cliente:
        registros = registros.filter(nome_cliente__nome__icontains=cliente)
    
    if status_os:
        registros = registros.filter(status_os__nome=status_os)
    
    # Paginação
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    start = (page - 1) * page_size
    end = start + page_size
    
    total = registros.count()
    registros_paginados = registros[start:end]
    
    logger.info(f"Relatórios: Total de registros encontrados: {total}")
    
    # Serializar dados
    dados = []
    for registro in registros_paginados:
        dados.append({
            'id': registro.id,
            'numero_os': registro.numero_os,
            'data_solicitacao': registro.data_solicitacao_os.strftime('%d/%m/%Y %H:%M') if registro.data_solicitacao_os else '',
            'cliente': registro.nome_cliente.nome if registro.nome_cliente else '',
            'status': registro.status_os.nome if registro.status_os else '',
            'descricao': registro.descricao_resumida or '',
            'usuario_criacao': registro.usuario.username if registro.usuario else '',
            'valor_total': float(registro.saldo_final) if registro.saldo_final and float(registro.saldo_final) > 0 else float(registro.soma_valores) if registro.soma_valores else 0,
            'created_at': registro.created_at.strftime('%d/%m/%Y %H:%M'),
        })
    
    response_data = {
        'registros': dados,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size
    }
    
    logger.info(f"Relatórios: Retornando {len(dados)} registros para página {page}")
    
    return Response(response_data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def relatorios_exportar_excel(request):
    """
    Exporta registros em formato Excel com formatação adequada (apenas para administradores e superiores)
    """
    try:
        logger.info(f"Iniciando exportação Excel para usuário: {request.user.username}")
        
        if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
            logger.warning(f"Tentativa de exportação por usuário sem permissão: {request.user.username}")
            return Response({
                'error': 'Acesso negado. Apenas administradores e superiores podem exportar relatórios.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Parâmetros de filtro
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        cliente = request.GET.get('cliente')
        status_os = request.GET.get('status_os')
        
        # Parâmetros de registros selecionados
        registros_selecionados = request.GET.getlist('registros_selecionados')
        
        logger.info(f"Filtros aplicados - data_inicio: {data_inicio}, data_fim: {data_fim}, cliente: {cliente}, status_os: {status_os}")
        logger.info(f"Registros selecionados: {registros_selecionados}")
        
        # Query base com todos os relacionamentos necessários
        registros = RegistroOS.objects.select_related(
            'nome_cliente', 'status_os', 'usuario', 'numero_contrato', 'unidade_cliente',
            'setor_unidade_cliente', 'status_regime_os', 'nome_diligenciador_os',
            'nome_solicitante_cliente', 'nome_responsavel_aprovacao_os_cliente',
            'nome_responsavel_execucao_servico', 'id_demanda', 'status_os_manual',
            'status_os_eletronica', 'status_levantamento', 'status_producao'
        ).prefetch_related(
            'documentos_solicitacao', 'datas_previstas', 'acoes_solicitacao',
            'controles_qualidade', 'ordens_cliente', 'documentos_entrada',
            'levantamentos', 'materiais', 'gmis', 'gmes', 'rtips', 'rtms',
            'dms', 'bms', 'frs', 'notas_fiscais_saida', 'notas_fiscais_venda'
        ).all()
        
        logger.info(f"Query base criada, total de registros: {registros.count()}")
        
        # Aplicar filtros
        if data_inicio:
            try:
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
                registros = registros.filter(created_at__gte=data_inicio)
                logger.info(f"Filtro data_inicio aplicado: {data_inicio}")
            except ValueError:
                logger.error(f"Formato de data de início inválido: {data_inicio}")
                return Response({
                    'error': 'Formato de data de início inválido. Use o formato YYYY-MM-DD.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if data_fim:
            try:
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                registros = registros.filter(created_at__lt=data_fim)
                logger.info(f"Filtro data_fim aplicado: {data_fim}")
            except ValueError:
                logger.error(f"Formato de data de fim inválido: {data_fim}")
                return Response({
                    'error': 'Formato de data de fim inválido. Use o formato YYYY-MM-DD.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if cliente:
            registros = registros.filter(nome_cliente__nome__icontains=cliente)
            logger.info(f"Filtro cliente aplicado: {cliente}")
        
        if status_os:
            registros = registros.filter(status_os__nome=status_os)
            logger.info(f"Filtro status_os aplicado: {status_os}")
        
        # Filtrar por registros selecionados se especificado
        if registros_selecionados:
            try:
                registros = registros.filter(id__in=registros_selecionados)
                logger.info(f"Filtro registros selecionados aplicado: {len(registros_selecionados)} registros")
            except ValueError:
                logger.error(f"IDs de registros selecionados inválidos: {registros_selecionados}")
                return Response({
                    'error': 'IDs de registros selecionados inválidos.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verificar se há registros para exportar
        total_registros = registros.count()
        logger.info(f"Total de registros após filtros: {total_registros}")
        
        if not registros.exists():
            logger.warning("Nenhum registro encontrado com os filtros aplicados")
            return Response({
                'error': 'Nenhum registro encontrado com os filtros aplicados.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        logger.info("Iniciando criação do workbook Excel")
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório OS"
        
        # Definir cabeçalhos organizados em blocos
        headers = [
            # Dados Gerais da OS
            'Número OS', 'Cliente', 'Status OS', 'Data Solicitação', 'Data Emissão',
            'Prazo Execução', 'Regime OS', 'Contrato', 'Unidade Cliente', 'Setor Unidade',
            'Diligenciador', 'Solicitante Cliente', 'Responsável Aprovação', 'Responsável Execução',
            'Demanda', 'Descrição Resumida', 'Descrição Detalhada', 'Existe Orçamento',
            
            # Medições
            'Peso Fabricação (kg)', 'Metro Quadrado Pintura (m²)',
            
            # Valores - Fabricação
            'Valor Fabricação (R$)', 'Haverá Valor Fabricação',
            'Valor Material Fabricação (R$)', 'Haverá Valor Material Fabricação',
            
            # Valores - Levantamento
            'Valor Levantamento (R$)', 'Haverá Valor Levantamento',
            
            # Valores - Pintura
            'Valor Material Pintura (R$)', 'Haverá Valor Material Pintura',
            'Valor Serviço Pintura (R$)', 'Haverá Valor Serviço Pintura',
            
            # Valores - Montagem
            'Valor Montagem (R$)', 'Haverá Valor Montagem',
            'Valor Material Montagem (R$)', 'Haverá Valor Material Montagem',
            
            # Valores - Outros
            'Valor Inspeção (R$)', 'Haverá Valor Inspeção',
            'Valor HH (R$)', 'Haverá Valor HH',
            'Valor Manutenção Válvula (R$)', 'Haverá Valor Manutenção Válvula',
            'Valor Serviço Terceiros (R$)', 'Haverá Valor Serviço Terceiros',
            
            # Totais
            'Soma Valores (R$)', 'HH Previsão', 'Soma Notas Fiscais (R$)', 'Saldo Final (R$)',
            
            # Status e Controle
            'Status OS Manual', 'Status OS Eletrônica', 'Data Aprovação Manual',
            'Data Assinatura Eletrônica', 'Número OS Eletrônica',
            'Status Levantamento', 'Status Produção',
            
            # Documentos
            'Opções DMS', 'Opções BMS', 'Opções FRS', 'Opções NF',
            
            # Observações e Controle
            'Observação', 'Usuário Criação', 'Data Criação', 'Data Atualização'
        ]
        
        logger.info("Adicionando cabeçalhos ao Excel")
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        logger.info("Processando dados dos registros")
        
        # Adicionar dados
        row_num = 2
        registros_processados = 0
        for registro in registros:
            try:
                col_num = 1
                
                # Dados Gerais da OS
                ws.cell(row=row_num, column=col_num, value=registro.numero_os); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.nome_cliente.nome if registro.nome_cliente else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.status_os.nome if registro.status_os else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.data_solicitacao_os.strftime('%d/%m/%Y %H:%M') if registro.data_solicitacao_os else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.data_emissao_os.strftime('%d/%m/%Y %H:%M') if registro.data_emissao_os else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.prazo_execucao_servico.strftime('%d/%m/%Y %H:%M') if registro.prazo_execucao_servico else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.status_regime_os.nome if registro.status_regime_os else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.numero_contrato.numero if registro.numero_contrato else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.unidade_cliente.nome if registro.unidade_cliente else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.setor_unidade_cliente.nome if registro.setor_unidade_cliente else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.nome_diligenciador_os.nome if registro.nome_diligenciador_os else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.nome_solicitante_cliente.nome if registro.nome_solicitante_cliente else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.nome_responsavel_aprovacao_os_cliente.nome if registro.nome_responsavel_aprovacao_os_cliente else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.nome_responsavel_execucao_servico.nome if registro.nome_responsavel_execucao_servico else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.id_demanda.nome if registro.id_demanda else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.descricao_resumida or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.descricao_detalhada or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.existe_orcamento or ''); col_num += 1
                
                # Medições
                ws.cell(row=row_num, column=col_num, value=float(registro.peso_fabricacao) if registro.peso_fabricacao else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.metro_quadrado_pintura_revestimento) if registro.metro_quadrado_pintura_revestimento else 0); col_num += 1
                
                # Valores - Fabricação
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_fabricacao) if registro.valor_fabricacao else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_fabricacao or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_material_fabricacao) if registro.valor_material_fabricacao else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_material_fabricacao or ''); col_num += 1
                
                # Valores - Levantamento
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_levantamento) if registro.valor_levantamento else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_levantamento or ''); col_num += 1
                
                # Valores - Pintura
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_material_pintura) if registro.valor_material_pintura else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_material_pintura or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_servico_pintura_revestimento) if registro.valor_servico_pintura_revestimento else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_servico_pintura_revestimento or ''); col_num += 1
                
                # Valores - Montagem
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_montagem) if registro.valor_montagem else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_montagem or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_material_montagem) if registro.valor_material_montagem else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_material_montagem or ''); col_num += 1
                
                # Valores - Outros
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_inspecao) if registro.valor_inspecao else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_inspecao or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_hh) if registro.valor_hh else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_hh or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_manutencao_valvula) if registro.valor_manutencao_valvula else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_manutencao_valvula or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.valor_servico_terceiros) if registro.valor_servico_terceiros else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.havera_valor_servico_terceiros or ''); col_num += 1
                
                # Totais
                ws.cell(row=row_num, column=col_num, value=float(registro.soma_valores) if registro.soma_valores else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.hh_previsao) if registro.hh_previsao else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.soma_notas_fiscais) if registro.soma_notas_fiscais else 0); col_num += 1
                ws.cell(row=row_num, column=col_num, value=float(registro.saldo_final) if registro.saldo_final else 0); col_num += 1
                
                # Status e Controle
                ws.cell(row=row_num, column=col_num, value=registro.status_os_manual.nome if registro.status_os_manual else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.status_os_eletronica.nome if registro.status_os_eletronica else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.data_aprovacao_assinatura_manual.strftime('%d/%m/%Y %H:%M') if registro.data_aprovacao_assinatura_manual else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.data_assinatura_eletronica_os.strftime('%d/%m/%Y %H:%M') if registro.data_assinatura_eletronica_os else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.numero_os_eletronica or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.status_levantamento.nome if registro.status_levantamento else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.status_producao.nome if registro.status_producao else ''); col_num += 1
                
                # Documentos
                ws.cell(row=row_num, column=col_num, value=registro.opcoes_dms or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.opcoes_bms or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.opcoes_frs or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.opcoes_nf or ''); col_num += 1
                
                # Observações e Controle
                ws.cell(row=row_num, column=col_num, value=registro.observacao or ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.usuario.username if registro.usuario else ''); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.created_at.strftime('%d/%m/%Y %H:%M')); col_num += 1
                ws.cell(row=row_num, column=col_num, value=registro.updated_at.strftime('%d/%m/%Y %H:%M')); col_num += 1
                
                # Aplicar formatação condicional
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Formatação para valores monetários
                    if 'R$' in headers[col-1]:
                        cell.number_format = 'R$ #,##0.00'
                        if cell.value and float(cell.value) > 0:
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    
                    # Formatação para status
                    if 'Status' in headers[col-1]:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    
                    # Formatação para saldo final
                    if 'Saldo Final' in headers[col-1]:
                        cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                        cell.font = Font(bold=True)
                
                row_num += 1
                registros_processados += 1
                
                # Log a cada 100 registros processados
                if registros_processados % 100 == 0:
                    logger.info(f"Processados {registros_processados} registros de {total_registros}")
                    
            except Exception as e:
                logger.error(f"Erro ao processar registro {registro.id}: {str(e)}")
                continue
        
        logger.info(f"Total de registros processados: {registros_processados}")
        
        logger.info("Ajustando largura das colunas")
        
        # Ajustar largura das colunas com largura mínima adequada
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        # Para cabeçalhos, considerar quebra de linha
                        if cell.row == 1:
                            # Contar o número de palavras para estimar largura
                            words = str(cell.value).split()
                            cell_length = max(len(word) for word in words) if words else cell_length
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Largura mínima baseada no tipo de coluna
            if 'Status' in headers[column[0].column - 1]:
                min_width = 15
            elif 'R$' in headers[column[0].column - 1]:
                min_width = 18
            elif 'Data' in headers[column[0].column - 1]:
                min_width = 20
            elif 'Descrição' in headers[column[0].column - 1]:
                min_width = 25
            elif 'Observação' in headers[column[0].column - 1]:
                min_width = 30
            else:
                min_width = 12
            
            adjusted_width = max(min_width, min(max_length + 3, 60))
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar filtros automáticos
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num-1}"
        
        logger.info("Criando aba de resumo")
        
        # Criar aba de resumo
        ws_resumo = wb.create_sheet("Resumo")
        
        # Estatísticas
        total_os = len(registros)
        soma_total_valores = sum(float(r.soma_valores) if r.soma_valores else 0 for r in registros)
        soma_total_notas_fiscais = sum(float(r.soma_notas_fiscais) if r.soma_notas_fiscais else 0 for r in registros)
        soma_total_saldo_final = sum(float(r.saldo_final) if r.saldo_final else 0 for r in registros)
        
        # Cabeçalho do resumo
        ws_resumo['A1'] = "RESUMO DO RELATÓRIO"
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_resumo['A1'].font = Font(bold=True, color="FFFFFF")
        
        # Dados do resumo
        ws_resumo['A3'] = "Total de Ordens de Serviço:"
        ws_resumo['B3'] = total_os
        ws_resumo['A4'] = "Soma Total dos Valores:"
        ws_resumo['B4'] = soma_total_valores
        ws_resumo['B4'].number_format = 'R$ #,##0.00'
        ws_resumo['A5'] = "Soma Total das Notas Fiscais:"
        ws_resumo['B5'] = soma_total_notas_fiscais
        ws_resumo['B5'].number_format = 'R$ #,##0.00'
        ws_resumo['A6'] = "Saldo Final Total:"
        ws_resumo['B6'] = soma_total_saldo_final
        ws_resumo['B6'].number_format = 'R$ #,##0.00'
        
        # Formatação do resumo
        for row in range(3, 7):
            ws_resumo[f'A{row}'].font = Font(bold=True)
            ws_resumo[f'B{row}'].font = Font(bold=True)
        
        # Ajustar largura das colunas do resumo
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20
        
        logger.info("Salvando arquivo Excel")
        
        # Salvar para buffer
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        logger.info(f"Arquivo Excel gerado com sucesso. Tamanho: {len(excel_file.getvalue())} bytes")
        
        # Criar resposta HTTP
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_os_formatado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        logger.info("Exportação Excel concluída com sucesso")
        return response
        
    except Exception as e:
        logger.error(f"Erro na exportação Excel: {str(e)}")
        import traceback
        logger.error(f"Traceback completo: {traceback.format_exc()}")
        return Response({
            'error': 'Erro interno na geração do relatório Excel. Tente novamente ou entre em contato com o suporte.'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def relatorios_exportar_pdf(request):
    """
    Exporta registros em formato PDF com layout estruturado (apenas para administradores e superiores)
    """
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem exportar relatórios.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Parâmetros de filtro
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    cliente = request.GET.get('cliente')
    status_os = request.GET.get('status_os')
    
    # Parâmetros de registros selecionados
    registros_selecionados = request.GET.getlist('registros_selecionados')
    
    # Query base com todos os relacionamentos necessários
    registros = RegistroOS.objects.select_related(
        'nome_cliente', 'status_os', 'usuario', 'numero_contrato', 'unidade_cliente',
        'setor_unidade_cliente', 'status_regime_os', 'nome_diligenciador_os',
        'nome_solicitante_cliente', 'nome_responsavel_aprovacao_os_cliente',
        'nome_responsavel_execucao_servico', 'id_demanda', 'status_os_manual',
        'status_os_eletronica', 'status_levantamento', 'status_producao'
    ).prefetch_related(
        'documentos_solicitacao', 'datas_previstas', 'acoes_solicitacao',
        'controles_qualidade', 'ordens_cliente', 'documentos_entrada',
        'levantamentos', 'materiais', 'gmis', 'gmes', 'rtips', 'rtms',
        'dms', 'bms', 'frs', 'notas_fiscais_saida', 'notas_fiscais_venda'
    ).all()
    
    # Aplicar filtros
    if data_inicio:
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            registros = registros.filter(created_at__gte=data_inicio)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            registros = registros.filter(created_at__lt=data_fim)
        except ValueError:
            pass
    
    if cliente:
        registros = registros.filter(nome_cliente__nome__icontains=cliente)
    
    if status_os:
        registros = registros.filter(status_os__nome=status_os)
    
    # Filtrar por registros selecionados se especificado
    if registros_selecionados:
        registros = registros.filter(id__in=registros_selecionados)
    
    # Criar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_os_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Criar documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=20,
        spaceBefore=20
    )
    
    # Título do relatório
    story.append(Paragraph("RELATÓRIO DE ORDENS DE SERVIÇO", title_style))
    story.append(Spacer(1, 20))
    
    # Informações do filtro
    filtros_info = []
    if data_inicio:
        filtros_info.append(f"Data Início: {data_inicio.strftime('%d/%m/%Y')}")
    if data_fim:
        filtros_info.append(f"Data Fim: {(data_fim - timedelta(days=1)).strftime('%d/%m/%Y')}")
    if cliente:
        filtros_info.append(f"Cliente: {cliente}")
    if status_os:
        filtros_info.append(f"Status: {status_os}")
    
    if filtros_info:
        story.append(Paragraph("Filtros Aplicados:", subtitle_style))
        for filtro in filtros_info:
            story.append(Paragraph(f"• {filtro}", styles['Normal']))
        story.append(Spacer(1, 20))
    
    # Estatísticas
    total_registros = registros.count()
    valor_total = sum(
        float(r.saldo_final) if r.saldo_final and float(r.saldo_final) > 0 
        else float(r.soma_valores) if r.soma_valores else 0 
        for r in registros
    )
    usuarios_unicos = len(set(r.usuario.username for r in registros if r.usuario))
    
    story.append(Paragraph("Estatísticas:", subtitle_style))
    story.append(Paragraph(f"• Total de Registros: {total_registros}", styles['Normal']))
    story.append(Paragraph(f"• Valor Total: R$ {valor_total:,.2f}", styles['Normal']))
    story.append(Paragraph(f"• Usuários Únicos: {usuarios_unicos}", styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Para cada registro, criar uma seção detalhada
    for i, registro in enumerate(registros, 1):
        story.append(Paragraph(f"ORDEM DE SERVIÇO #{registro.numero_os}", subtitle_style))
        
        # Dados básicos
        dados_basicos = [
            ['Campo', 'Valor'],
            ['Número OS', str(registro.numero_os)],
            ['Data Solicitação', registro.data_solicitacao_os.strftime('%d/%m/%Y %H:%M') if registro.data_solicitacao_os else 'N/A'],
            ['Data Emissão', registro.data_emissao_os.strftime('%d/%m/%Y %H:%M') if registro.data_emissao_os else 'N/A'],
            ['Cliente', registro.nome_cliente.nome if registro.nome_cliente else 'N/A'],
            ['Contrato', registro.numero_contrato.numero if registro.numero_contrato else 'N/A'],
            ['Unidade Cliente', registro.unidade_cliente.nome if registro.unidade_cliente else 'N/A'],
            ['Setor Unidade', registro.setor_unidade_cliente.nome if registro.setor_unidade_cliente else 'N/A'],
            ['Prazo Execução', registro.prazo_execucao_servico.strftime('%d/%m/%Y %H:%M') if registro.prazo_execucao_servico else 'N/A'],
            ['Regime OS', registro.status_regime_os.nome if registro.status_regime_os else 'N/A'],
            ['Diligenciador', registro.nome_diligenciador_os.nome if registro.nome_diligenciador_os else 'N/A'],
            ['Solicitante Cliente', registro.nome_solicitante_cliente.nome if registro.nome_solicitante_cliente else 'N/A'],
            ['Responsável Aprovação', registro.nome_responsavel_aprovacao_os_cliente.nome if registro.nome_responsavel_aprovacao_os_cliente else 'N/A'],
            ['Responsável Execução', registro.nome_responsavel_execucao_servico.nome if registro.nome_responsavel_execucao_servico else 'N/A'],
            ['Demanda', registro.id_demanda.nome if registro.id_demanda else 'N/A'],
            ['Status', registro.status_os.nome if registro.status_os else 'N/A'],
            ['Status Manual', registro.status_os_manual.nome if registro.status_os_manual else 'N/A'],
            ['Status Eletrônica', registro.status_os_eletronica.nome if registro.status_os_eletronica else 'N/A'],
            ['Status Levantamento', registro.status_levantamento.nome if registro.status_levantamento else 'N/A'],
            ['Status Produção', registro.status_producao.nome if registro.status_producao else 'N/A'],
            ['Usuário Criação', registro.usuario.username if registro.usuario else 'N/A'],
        ]
        
        t1 = Table(dados_basicos, colWidths=[2*inch, 4*inch])
        t1.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(t1)
        story.append(Spacer(1, 15))
        
        # Descrições
        if registro.descricao_resumida or registro.descricao_detalhada:
            story.append(Paragraph("Descrições:", styles['Heading3']))
            if registro.descricao_resumida:
                story.append(Paragraph(f"<b>Resumida:</b> {registro.descricao_resumida}", styles['Normal']))
            if registro.descricao_detalhada:
                story.append(Paragraph(f"<b>Detalhada:</b> {registro.descricao_detalhada}", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Valores
        story.append(Paragraph("Valores:", styles['Heading3']))
        valores_data = [
            ['Tipo', 'Valor', 'Incluir'],
            ['Fabricação', f"R$ {float(registro.valor_fabricacao or 0):,.2f}", registro.havera_valor_fabricacao or 'N/A'],
            ['Levantamento', f"R$ {float(registro.valor_levantamento or 0):,.2f}", registro.havera_valor_levantamento or 'N/A'],
            ['Material Fabricação', f"R$ {float(registro.valor_material_fabricacao or 0):,.2f}", registro.havera_valor_material_fabricacao or 'N/A'],
            ['Material Pintura', f"R$ {float(registro.valor_material_pintura or 0):,.2f}", registro.havera_valor_material_pintura or 'N/A'],
            ['Serviço Pintura', f"R$ {float(registro.valor_servico_pintura_revestimento or 0):,.2f}", registro.havera_valor_servico_pintura_revestimento or 'N/A'],
            ['Montagem', f"R$ {float(registro.valor_montagem or 0):,.2f}", registro.havera_valor_montagem or 'N/A'],
            ['Material Montagem', f"R$ {float(registro.valor_material_montagem or 0):,.2f}", registro.havera_valor_material_montagem or 'N/A'],
            ['Inspeção', f"R$ {float(registro.valor_inspecao or 0):,.2f}", registro.havera_valor_inspecao or 'N/A'],
            ['HH', f"R$ {float(registro.valor_hh or 0):,.2f}", registro.havera_valor_hh or 'N/A'],
            ['Manutenção Válvula', f"R$ {float(registro.valor_manutencao_valvula or 0):,.2f}", registro.havera_valor_manutencao_valvula or 'N/A'],
            ['Serviço Terceiros', f"R$ {float(registro.valor_servico_terceiros or 0):,.2f}", registro.havera_valor_servico_terceiros or 'N/A'],
            ['<b>TOTAL</b>', f"<b>R$ {float(registro.soma_valores or 0):,.2f}</b>", ''],
        ]
        
        # Adicionar informações de medições
        if registro.peso_fabricacao or registro.metro_quadrado_pintura_revestimento:
            story.append(Paragraph("Medições:", styles['Heading3']))
            if registro.peso_fabricacao:
                story.append(Paragraph(f"<b>Peso Fabricação:</b> {float(registro.peso_fabricacao):,.2f} kg", styles['Normal']))
            if registro.metro_quadrado_pintura_revestimento:
                story.append(Paragraph(f"<b>Metro Quadrado Pintura:</b> {float(registro.metro_quadrado_pintura_revestimento):,.2f} m²", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Adicionar informações de documentos
        if registro.opcoes_dms or registro.opcoes_bms or registro.opcoes_frs or registro.opcoes_nf:
            story.append(Paragraph("Documentos:", styles['Heading3']))
            if registro.opcoes_dms:
                story.append(Paragraph(f"<b>DMS:</b> {registro.opcoes_dms}", styles['Normal']))
            if registro.opcoes_bms:
                story.append(Paragraph(f"<b>BMS:</b> {registro.opcoes_bms}", styles['Normal']))
            if registro.opcoes_frs:
                story.append(Paragraph(f"<b>FRS:</b> {registro.opcoes_frs}", styles['Normal']))
            if registro.opcoes_nf:
                story.append(Paragraph(f"<b>NF:</b> {registro.opcoes_nf}", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Adicionar informações de controle
        if registro.data_aprovacao_assinatura_manual or registro.data_assinatura_eletronica_os or registro.numero_os_eletronica:
            story.append(Paragraph("Controle:", styles['Heading3']))
            if registro.data_aprovacao_assinatura_manual:
                story.append(Paragraph(f"<b>Data Aprovação Manual:</b> {registro.data_aprovacao_assinatura_manual.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            if registro.data_assinatura_eletronica_os:
                story.append(Paragraph(f"<b>Data Assinatura Eletrônica:</b> {registro.data_assinatura_eletronica_os.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            if registro.numero_os_eletronica:
                story.append(Paragraph(f"<b>Número OS Eletrônica:</b> {registro.numero_os_eletronica}", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Adicionar informações de totais
        story.append(Paragraph("Totais:", styles['Heading3']))
        story.append(Paragraph(f"<b>Soma Valores:</b> R$ {float(registro.soma_valores or 0):,.2f}", styles['Normal']))
        story.append(Paragraph(f"<b>HH Previsão:</b> {float(registro.hh_previsao or 0):,.2f}", styles['Normal']))
        story.append(Paragraph(f"<b>Soma Notas Fiscais:</b> R$ {float(registro.soma_notas_fiscais or 0):,.2f}", styles['Normal']))
        story.append(Paragraph(f"<b>Saldo Final:</b> R$ {float(registro.saldo_final or 0):,.2f}", styles['Normal']))
        story.append(Spacer(1, 15))
        
        t2 = Table(valores_data, colWidths=[2*inch, 1.5*inch, 1*inch])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(t2)
        story.append(Spacer(1, 20))
        
        # Observação se houver
        if registro.observacao:
            story.append(Paragraph("Observação:", styles['Heading3']))
            story.append(Paragraph(registro.observacao, styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Quebra de página se não for o último
        if i < len(registros):
            story.append(PageBreak())
    
    # Construir PDF
    doc.build(story)
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def configuracoes_listar_usuarios(request):
    """
    Lista todos os usuários para configurações (apenas para administradores e superiores)
    """
    logger.info(f"Configurações: Usuário {request.user.username} tentando acessar configurações")
    
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        logger.warning(f"Configurações: Acesso negado para usuário {request.user.username}")
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem acessar configurações.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    usuarios = User.objects.all().order_by('username')
    dados = []
    
    for usuario in usuarios:
        grupos = [{'id': g.id, 'name': g.name} for g in usuario.groups.all()]
        dados.append({
            'id': usuario.id,
            'username': usuario.username,
            'email': usuario.email,
            'first_name': usuario.first_name,
            'last_name': usuario.last_name,
            'is_active': usuario.is_active,
            'date_joined': usuario.date_joined.strftime('%d/%m/%Y %H:%M'),
            'groups': grupos,
            'is_staff': usuario.is_staff,
            'is_superuser': usuario.is_superuser,
        })
    
    logger.info(f"Configurações: Retornando {len(dados)} usuários")
    
    return Response(dados)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def configuracoes_alterar_grupo_usuario(request):
    """
    Altera o grupo de um usuário (apenas para administradores e superiores)
    """
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem alterar grupos.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    user_id = request.data.get('user_id')
    group_ids = request.data.get('group_ids', [])
    
    if not user_id:
        return Response({
            'error': 'ID do usuário é obrigatório'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({
            'error': 'Usuário não encontrado'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Limpar grupos atuais
    usuario.groups.clear()
    
    # Adicionar novos grupos
    for group_id in group_ids:
        try:
            grupo = Group.objects.get(id=group_id)
            usuario.groups.add(grupo)
        except Group.DoesNotExist:
            logger.warning(f"Grupo não encontrado: {group_id}")
    
    logger.info(f"Grupos alterados para usuário {usuario.username} por {request.user.username}")
    
    # Retornar dados atualizados
    grupos = [{'id': g.id, 'name': g.name} for g in usuario.groups.all()]
    return Response({
        'message': 'Grupos alterados com sucesso',
        'user': {
            'id': usuario.id,
            'username': usuario.username,
            'groups': grupos
        }
    })


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def cadastro_usuario(request):
    """
    Cadastro de novos usuários (público)
    """
    username = request.data.get('username')
    email = request.data.get('email')
    password = request.data.get('password')
    first_name = request.data.get('first_name', '')
    last_name = request.data.get('last_name', '')
    
    if not all([username, email, password]):
        return Response({
            'error': 'Username, email e password são obrigatórios'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if len(password) < 8:
        return Response({
            'error': 'A senha deve ter pelo menos 8 caracteres'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if User.objects.filter(username=username).exists():
        return Response({
            'error': 'Username já existe'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if User.objects.filter(email=email).exists():
        return Response({
            'error': 'Email já existe'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Criar usuário inativo (aguardando aprovação)
    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
        is_active=False  # Usuário inativo até ser aprovado
    )
    
    logger.info(f"Novo usuário cadastrado: {username} (aguardando aprovação)")
    
    return Response({
        'message': 'Cadastro realizado com sucesso. Aguarde a aprovação de um administrador.',
        'user_id': user.id
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def configuracoes_ativar_usuario(request):
    """
    Ativa um usuário (apenas para administradores e superiores)
    """
    if not request.user.groups.filter(name__in=['Administrador', 'Superior']).exists():
        return Response({
            'error': 'Acesso negado. Apenas administradores e superiores podem ativar usuários.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    user_id = request.data.get('user_id')
    is_active = request.data.get('is_active', True)
    
    if not user_id:
        return Response({
            'error': 'ID do usuário é obrigatório'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({
            'error': 'Usuário não encontrado'
        }, status=status.HTTP_404_NOT_FOUND)
    
    usuario.is_active = is_active
    usuario.save()
    
    logger.info(f"Usuário {usuario.username} {'ativado' if is_active else 'desativado'} por {request.user.username}")
    
    return Response({
        'message': f'Usuário {"ativado" if is_active else "desativado"} com sucesso',
        'user': {
            'id': usuario.id,
            'username': usuario.username,
            'is_active': usuario.is_active
        }
    })


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated])
def perfil_atualizar(request):
    """
    Permite ao usuário atualizar seus próprios dados
    """
    logger.info(f"Perfil: Usuário {request.user.username} tentando atualizar perfil")
    
    user = request.user
    
    # Campos permitidos para atualização
    first_name = request.data.get('first_name')
    last_name = request.data.get('last_name')
    email = request.data.get('email')
    
    logger.info(f"Perfil: Dados recebidos - first_name={first_name}, last_name={last_name}, email={email}")
    
    if first_name is not None:
        user.first_name = first_name
    
    if last_name is not None:
        user.last_name = last_name
    
    if email is not None:
        # Verificar se email já existe
        if User.objects.filter(email=email).exclude(id=user.id).exists():
            logger.warning(f"Perfil: Email já existe: {email}")
            return Response({
                'error': 'Email já está em uso por outro usuário'
            }, status=status.HTTP_400_BAD_REQUEST)
        user.email = email
    
    user.save()
    
    logger.info(f"Perfil: Dados do perfil atualizados para o usuário: {user.username}")
    
    # Retornar dados atualizados
    user_serializer = UserSerializer(user)
    groups = [group.name for group in user.groups.all()]
    
    return Response({
        'message': 'Perfil atualizado com sucesso',
        'user': user_serializer.data,
        'groups': groups
    })


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def configuracoes_excluir_usuario(request):
    """
    Exclui um usuário (apenas para administradores)
    """
    if not request.user.groups.filter(name='Administrador').exists():
        return Response({
            'error': 'Acesso negado. Apenas administradores podem excluir usuários.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    user_id = request.data.get('user_id')
    
    if not user_id:
        return Response({
            'error': 'ID do usuário é obrigatório'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({
            'error': 'Usuário não encontrado'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Verificar se o usuário não está tentando excluir a si mesmo
    if usuario.id == request.user.id:
        return Response({
            'error': 'Você não pode excluir sua própria conta'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Verificar se o usuário não é superusuário
    if usuario.is_superuser:
        return Response({
            'error': 'Não é possível excluir um superusuário'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    username = usuario.username
    usuario.delete()
    
    logger.info(f"Usuário {username} excluído por {request.user.username}")
    
    return Response({
        'message': f'Usuário {username} excluído com sucesso'
    })

